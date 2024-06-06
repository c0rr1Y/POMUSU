#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import time

from config.data import File
import openpyxl


class output:
    def __init__(self):
        # 获取当前时间并格式化为字符串
        self.nowTime = time.strftime("%Y%m%d%H%M%S", time.localtime())
        self.filename_xls = f"{self.nowTime}.xlsx"
        # 定义输出目录
        self.output_dir = 'output'
        # 确保输出目录存在
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        # 构建保存文件的完整路径
        self.path_xls = os.path.join(self.output_dir, self.filename_xls)
        self.write()
    def write(self):

        # 创建一个新的Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vulnerability Data"
        # 写入表头
        headers = ['URL', 'CMS', 'Title', 'Status', 'Server', 'Vulnerability']
        ws.append(headers)
        # 写入数据
        for data in File.file:
            vul_str = ', '.join(data['vul'])  # 将vul列表转换为逗号分隔的字符串
            ws.append([data['url'], data['cms'], data['title'], data['status'], data['Server'], vul_str])
        vul_column_letter = openpyxl.utils.get_column_letter(headers.index('Vulnerability') + 1)
        URL = openpyxl.utils.get_column_letter(headers.index('URL') + 1)
        CMS = openpyxl.utils.get_column_letter(headers.index('CMS') + 1)
        ws.column_dimensions[URL].width = 30
        ws.column_dimensions[CMS].width = 30
        ws.column_dimensions[vul_column_letter].width = 100
        # 保存Excel文件
        wb.save(self.path_xls)
